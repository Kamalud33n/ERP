import io
from datetime import datetime
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment


# ── Shared PDF Builder ─────────────────────────────────
def build_pdf_report(title: str, subtitle: str, headers: list, rows: list, summary: dict = None) -> io.BytesIO:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=2*cm, leftMargin=2*cm, topMargin=2*cm, bottomMargin=2*cm)
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle('Title', parent=styles['Title'], fontSize=20, textColor=colors.HexColor('#1e293b'), alignment=TA_CENTER)
    sub_style   = ParagraphStyle('Sub', parent=styles['Normal'], fontSize=11, textColor=colors.HexColor('#64748b'), alignment=TA_CENTER, spaceAfter=10)
    date_style  = ParagraphStyle('Date', parent=styles['Normal'], fontSize=9, textColor=colors.HexColor('#94a3b8'), alignment=TA_CENTER, spaceAfter=20)

    story = []
    story.append(Paragraph(title, title_style))
    story.append(Paragraph(subtitle, sub_style))
    story.append(Paragraph(f"Generated on {datetime.now().strftime('%Y-%m-%d %H:%M')}", date_style))
    story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor('#2563eb')))
    story.append(Spacer(1, 0.4*cm))

    # Summary cards
    if summary:
        summary_data = [[k, str(v)] for k, v in summary.items()]
        s_table = Table(summary_data, colWidths=[8*cm, 8*cm])
        s_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (0,-1), colors.HexColor('#f1f5f9')),
            ('FONTNAME',   (0,0), (0,-1), 'Helvetica-Bold'),
            ('FONTSIZE',   (0,0), (-1,-1), 10),
            ('GRID',       (0,0), (-1,-1), 0.5, colors.HexColor('#e2e8f0')),
            ('PADDING',    (0,0), (-1,-1), 8),
        ]))
        story.append(s_table)
        story.append(Spacer(1, 0.5*cm))

    # Main table
    table_data = [headers] + rows
    col_width = (17*cm) / len(headers)
    t = Table(table_data, colWidths=[col_width]*len(headers), repeatRows=1)
    t.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#2563eb')),
        ('TEXTCOLOR',  (0,0), (-1,0), colors.white),
        ('FONTNAME',   (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE',   (0,0), (-1,-1), 8),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#f8fafc')]),
        ('GRID',       (0,0), (-1,-1), 0.5, colors.HexColor('#e2e8f0')),
        ('PADDING',    (0,0), (-1,-1), 6),
        ('VALIGN',     (0,0), (-1,-1), 'MIDDLE'),
    ]))
    story.append(t)

    story.append(Spacer(1, 0.5*cm))
    story.append(HRFlowable(width="100%", thickness=0.5, color=colors.HexColor('#e2e8f0')))
    story.append(Paragraph("ERP System — Confidential Report", ParagraphStyle('Footer', parent=styles['Normal'], fontSize=8, textColor=colors.HexColor('#94a3b8'), alignment=TA_CENTER, spaceBefore=6)))

    doc.build(story)
    buffer.seek(0)
    return buffer


# ── Shared Excel Builder ───────────────────────────────
def build_excel_report(title: str, headers: list, rows: list, sheet_name: str = "Report") -> io.BytesIO:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Title row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.font = Font(size=14, bold=True, color="2563EB")
    title_cell.alignment = Alignment(horizontal="center")

    # Generated date
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    date_cell = ws.cell(row=2, column=1, value=f"Generated on {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    date_cell.font = Font(size=9, italic=True, color="94A3B8")
    date_cell.alignment = Alignment(horizontal="center")

    # Header row
    header_row = 4
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for r, row in enumerate(rows, start=header_row + 1):
        for c, value in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=value)
            if r % 2 == 0:
                cell.fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")

    # Auto column width
    for col in range(1, len(headers) + 1):
        max_len = max(
            [len(str(headers[col-1]))] +
            [len(str(row[col-1])) for row in rows] if rows else [10]
        )
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = max(max_len + 2, 12)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


# ── Payroll Export ─────────────────────────────────────
def export_payroll_pdf(payrolls: list) -> io.BytesIO:
    headers = ["ID", "Employee ID", "Month", "Basic", "Allowances", "Deductions", "Net Salary", "Status"]
    rows = [[p.id, p.employee_id, p.month, p.basic_salary, p.allowances, p.deductions, p.net_salary, p.status] for p in payrolls]
    total_net = sum(p.net_salary for p in payrolls)
    summary = {"Total Records": len(payrolls), "Total Net Salary": round(total_net, 2)}
    return build_pdf_report("Payroll Report", "Employee Salary Records", headers, rows, summary)

def export_payroll_excel(payrolls: list) -> io.BytesIO:
    headers = ["ID", "Employee ID", "Month", "Basic", "Allowances", "Deductions", "Net Salary", "Status"]
    rows = [[p.id, p.employee_id, p.month, p.basic_salary, p.allowances, p.deductions, p.net_salary, p.status] for p in payrolls]
    return build_excel_report("Payroll Report", headers, rows, "Payroll")


# ── Expense Export ──────────────────────────────────────
def export_expense_pdf(expenses: list) -> io.BytesIO:
    headers = ["ID", "Title", "Amount", "Category", "Department", "Date", "Status"]
    rows = [[e.id, e.title, e.amount, e.category, e.department, str(e.date), e.status] for e in expenses]
    total = sum(e.amount for e in expenses if e.status == "approved")
    summary = {"Total Records": len(expenses), "Total Approved Amount": round(total, 2)}
    return build_pdf_report("Expense Report", "Company Expense Records", headers, rows, summary)

def export_expense_excel(expenses: list) -> io.BytesIO:
    headers = ["ID", "Title", "Amount", "Category", "Department", "Date", "Status"]
    rows = [[e.id, e.title, e.amount, e.category, e.department, str(e.date), e.status] for e in expenses]
    return build_excel_report("Expense Report", headers, rows, "Expenses")


# ── Employee List Export ────────────────────────────────
def export_employees_excel(employees: list) -> io.BytesIO:
    headers = ["ID", "First Name", "Last Name", "Department", "Designation", "Salary", "Join Date"]
    rows = [[e.id, e.first_name, e.last_name, e.department, e.designation, e.salary, str(e.join_date) if e.join_date else "-"] for e in employees]
    return build_excel_report("Employee List", headers, rows, "Employees")

def export_employees_pdf(employees: list) -> io.BytesIO:
    headers = ["ID", "Name", "Department", "Designation", "Salary", "Join Date"]
    rows = [[e.id, f"{e.first_name} {e.last_name}", e.department, e.designation, e.salary, str(e.join_date) if e.join_date else "-"] for e in employees]
    summary = {"Total Employees": len(employees)}
    return build_pdf_report("Employee List", "All Employees", headers, rows, summary)


# ── Inventory Export ─────────────────────────────────────
def export_inventory_excel(items: list) -> io.BytesIO:
    headers = ["ID", "Name", "SKU", "Category", "Unit", "Current Stock", "Min Stock", "Unit Price", "Total Value"]
    rows = [[i.id, i.name, i.sku, i.category, i.unit, i.current_stock, i.min_stock, i.unit_price, round(i.current_stock * i.unit_price, 2)] for i in items]
    return build_excel_report("Inventory Report", headers, rows, "Inventory")


# ── Asset Export ──────────────────────────────────────────
def export_assets_excel(assets: list) -> io.BytesIO:
    headers = ["Code", "Name", "Category", "Location", "Department", "Purchase Price", "Current Value", "Status"]
    rows = [[a.asset_code, a.name, a.category, a.location or "-", a.department or "-", a.purchase_price, a.current_value, a.status] for a in assets]
    return build_excel_report("Asset Report", headers, rows, "Assets")


# ── P&L Export ────────────────────────────────────────────
def export_pl_pdf(pl_data: dict) -> io.BytesIO:
    headers = ["Item", "Amount"]
    rows = [
        ["Total Income",    pl_data["total_income"]],
        ["Total Expenses",  pl_data["total_expenses"]],
        ["Gross Profit",    pl_data["gross_profit"]],
        ["Net Profit",      pl_data["net_profit"]],
        ["Total Assets",    pl_data["total_assets"]],
        ["Total Liability", pl_data["total_liability"]],
    ]
    summary = {
        "Period": pl_data["period"],
        "Result": "Profitable" if pl_data["is_profitable"] else "Loss Making"
    }
    return build_pdf_report("Profit & Loss Report", f"Period: {pl_data['period']}", headers, rows, summary)